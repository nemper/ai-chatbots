import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

file = "data.xlsx"
df = pd.read_excel(file)

df = df.iloc[5:]    # testiranje

specific_column = "Ime ustanove:\n"
mapping_table = {"Da": 10, "Ne": 0, "Delimično tačno": 5, "Nisam siguran": 1, "Nemamo AV rešenje": 10}

matching_columns = df.columns[df.iloc[0].str.contains('|'.join(mapping_table.keys()), na=False)]

for col in matching_columns:
    df[col] = df[col].map(mapping_table)

keys1 = df[specific_column].to_list()
df = df[list(matching_columns)]
keys2 = df.columns.to_list()
values1 = df.sum(axis=1).to_list()
values2 = df.mean().round(2).to_list()

bodovanje_po_instituciji = dict(zip(keys1, values1))
bodovanje_po_pitanju = {k.rstrip('\n'): v for k, v in dict(zip(keys2, values2)).items()}


df1 = pd.DataFrame(list(bodovanje_po_instituciji.items()), columns=["Institucija", "Ukupno bodova"])
df1["Sredina"] = ""
df1[""] = ""
df1.sort_values(by=["Ukupno bodova"], inplace=True, ascending=False)

df2 = pd.DataFrame(list(bodovanje_po_pitanju.items()), columns=["Pitanje", "Prosečno bodova"])
df2.sort_values(by=["Prosečno bodova"], inplace=True, ascending=False, na_position="last")
frames = {"Institucija": df1, "Pitanje": df2}
df = pd.concat(frames)
x = len(df1)
df.loc[x:, ["Pitanje", "Prosečno bodova"]] = df.loc[:-x, ["Pitanje", "Prosečno bodova"]].values
print(df)
df.to_excel("statistika_upitnika.xlsx", index=False)

# boja
min_in_theory = 0
max_in_theory = 10 * 49

def gradient(value, min_value, max_value):
    normalized = (value - min_value) / (max_value - min_value)

    if normalized < 0.5:  # From dark red to yellow
        red = 255
        green = normalized * 2 * 255
        blue = 0
    else:  # From yellow to dark green
        red = (1 - normalized) * 2 * 255
        green = 255
        blue = 0  # Keep blue at 0
    
    return "{:02X}{:02X}{:02X}".format(int(red), int(green), int(blue))

wb = openpyxl.load_workbook("statistika_upitnika.xlsx")
sheet = wb.active

min_value = min(row[1].value for row in sheet.iter_rows() if isinstance(row[1].value, int))
max_value = max(row[1].value for row in sheet.iter_rows() if isinstance(row[1].value, int))

for row in sheet.iter_rows():
    if isinstance(row[1].value, int):
        fill_color = gradient(row[1].value, min_in_theory, max_in_theory)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        row[1].fill = fill
    if isinstance(row[5].value, int):
        fill_color = gradient(row[5].value, 0, 10)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        row[5].fill = fill
        fill_color = gradient(round(max_value / 2), min_value, max_value)
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        row[5].fill = fill

wb.save("statistika_upitnika.xlsx")

